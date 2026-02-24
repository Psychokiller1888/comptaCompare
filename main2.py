import json
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, date, timedelta
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Optional, Tuple

from openpyxl import load_workbook


# -----------------------------
# Config
# -----------------------------
abacusExportDir = Path('abacusExports')
bankExportDir = Path('bankExports')
logsDir = Path('logs')

WINDOW_DAYS = 2
DEC_Q = Decimal('0.01')


# -----------------------------
# Helpers
# -----------------------------
def q2(x: Decimal) -> Decimal:
	return x.quantize(DEC_Q, rounding=ROUND_HALF_UP)


def toDecimal(val: Any) -> Optional[Decimal]:
	if val is None:
		return None
	try:
		s = str(val).strip().replace("'", '').replace('’', '')
		if s == '':
			return None
		return q2(Decimal(s))
	except Exception:
		return None


def isFloat(string: str) -> bool:
	try:
		Decimal(str(string))
		return True
	except Exception:
		return False


def toDateStr(d: Any) -> Optional[str]:
	if d is None:
		return None

	if isinstance(d, str):
		ds = d.strip()
		for fmt in ('%d.%m.%Y', '%d.%m.%y', '%Y-%m-%d', '%d/%m/%Y'):
			try:
				parsed = datetime.strptime(ds, fmt).date()
				return parsed.strftime('%d.%m.%Y')
			except Exception:
				pass
		return ds

	if isinstance(d, datetime):
		return d.date().strftime('%d.%m.%Y')
	if isinstance(d, date):
		return d.strftime('%d.%m.%Y')

	return None


def parseDateStr(ds: str) -> date:
	return datetime.strptime(ds, '%d.%m.%Y').date()


def ensureLogsDir() -> None:
	logsDir.mkdir(parents=True, exist_ok=True)


def dumpJson(path: Path, obj: Any) -> None:
	ensureLogsDir()
	path.write_text(json.dumps(obj, ensure_ascii=False, indent='\t'), encoding='utf-8')


def chooseFile(stemChoices: list[str], title: str) -> str:
	print(f'\n{title}')
	for i, name in enumerate(stemChoices, start=1):
		print(f'  {i}) {name}')

	while True:
		raw = input('Choose number: ').strip()
		if not raw.isdigit():
			print('Please enter a number.')
			continue
		idx = int(raw)
		if 1 <= idx <= len(stemChoices):
			return stemChoices[idx - 1]
		print('Out of range.')


def askConfirm(message: str, default: bool = False) -> bool:
	suffix = 'Y/n' if default else 'y/N'
	while True:
		raw = input(f'{message} ({suffix}): ').strip().lower()
		if raw == '':
			return default
		if raw in ('y', 'yes', 'o', 'oui'):
			return True
		if raw in ('n', 'no', 'non'):
			return False
		print('Please answer y/n.')


def askFloat(message: str) -> Decimal:
	while True:
		raw = input(f'{message}: ').strip()
		if isFloat(raw):
			return q2(Decimal(raw))
		print('Please enter a valid number (example: -12.35).')


def buildDateCandidates(bookedAt: Any, valutaDate: Any, windowDays: int = WINDOW_DAYS) -> list[str]:
	# Priority: valuta first, then booked
	base: list[date] = []

	v = None
	b = None

	if isinstance(valutaDate, datetime):
		v = valutaDate.date()
	elif isinstance(valutaDate, date):
		v = valutaDate

	if isinstance(bookedAt, datetime):
		b = bookedAt.date()
	elif isinstance(bookedAt, date):
		b = bookedAt

	if v:
		base.append(v)
	if b and (not v or b != v):
		base.append(b)

	candidates: list[date] = []
	for d in base:
		candidates.append(d)
		for i in range(1, windowDays + 1):
			candidates.append(d + timedelta(days=i))
			candidates.append(d - timedelta(days=i))

	out: list[str] = []
	seen: set[str] = set()
	for d in candidates:
		ds = d.strftime('%d.%m.%Y')
		if ds not in seen:
			seen.add(ds)
			out.append(ds)
	return out


# -----------------------------
# Data Structures
# -----------------------------
Key = Tuple[str, Decimal]  # (dateStr, amount)


@dataclass
class BankTx:
	bookedAt: Optional[str]
	valutaDate: Optional[str]
	amount: Decimal
	isDebit: bool
	text: str
	balance: Optional[Decimal]


# -----------------------------
# Parse Abacus
# -----------------------------
def readAbacus(path: Path, rangeMin: date, rangeMax: date) -> dict:
	workbook = load_workbook(filename=str(path), data_only=True)
	sheet = workbook.active

	creditsCount: dict[Key, int] = defaultdict(int)
	debitsCount: dict[Key, int] = defaultdict(int)

	creditsByAmount: dict[Decimal, list[str]] = defaultdict(list)
	debitsByAmount: dict[Decimal, list[str]] = defaultdict(list)

	lines = 0

	startBalanceRange: Optional[Decimal] = None
	endBalanceRange: Optional[Decimal] = None

	for row in sheet.iter_rows(values_only=True):
		row0 = row[0]
		if row0 is None:
			continue

		row0s = str(row0).strip()

		if row0s.startswith('Solde'):
			continue

		try:
			dt = datetime.strptime(row0s, '%d.%m.%Y').date()
		except Exception:
			continue

		if dt < rangeMin or dt > rangeMax:
			continue

		ds = dt.strftime('%d.%m.%Y')

		debit = toDecimal(row[6] if len(row) > 6 else None)
		credit = toDecimal(row[7] if len(row) > 7 else None)
		bal = toDecimal(row[8] if len(row) > 8 else None)

		# Compute balance at range start (before first tx in range)
		# NOTE: In this Abacus export, 'Débit' increases the balance (inflow),
		# and 'Crédit' decreases the balance (outflow).
		if bal is not None:
			lineStartBalance = None
			if debit is not None and debit != 0:
				# debit is inflow -> previous balance = bal - debit
				lineStartBalance = q2(bal - debit)
			elif credit is not None and credit != 0:
				# credit is outflow -> previous balance = bal + credit
				lineStartBalance = q2(bal + credit)

			if startBalanceRange is None and lineStartBalance is not None:
				startBalanceRange = lineStartBalance

			endBalanceRange = bal

		if debit is not None and debit != 0:
			key = (ds, debit)
			debitsCount[key] += 1
			debitsByAmount[debit].append(ds)
			lines += 1
			continue

		if credit is not None and credit != 0:
			key = (ds, credit)
			creditsCount[key] += 1
			creditsByAmount[credit].append(ds)
			lines += 1
			continue

	if startBalanceRange is None:
		startBalanceRange = q2(Decimal('0'))
	if endBalanceRange is None:
		endBalanceRange = q2(Decimal('0'))

	return {
		'startBalance': q2(startBalanceRange),
		'endBalance': q2(endBalanceRange),
		'creditsCount': creditsCount,
		'debitsCount': debitsCount,
		'creditsByAmount': creditsByAmount,
		'debitsByAmount': debitsByAmount,
		'lines': lines,
		'rangeMin': rangeMin.strftime('%d.%m.%Y'),
		'rangeMax': rangeMax.strftime('%d.%m.%Y')
	}


# -----------------------------
# Parse Raiffeisen
# -----------------------------
def readRaiffeisen(path: Path) -> dict:
	workbook = load_workbook(filename=str(path), data_only=True)
	sheet = workbook.active

	txs: list[BankTx] = []
	endBalance: Optional[Decimal] = None
	lines = 0

	for row in sheet.iter_rows(values_only=True):
		if not row or len(row) < 6:
			continue

		iban, bookedAt, text, amount, balance, valutaDate = row[0], row[1], row[2], row[3], row[4], row[5]

		if isinstance(iban, str) and iban.lower().strip() == 'iban':
			continue

		if not iban:
			continue

		amt = toDecimal(amount)
		if amt is None or amt == 0:
			continue

		bal = toDecimal(balance)
		if bal is not None:
			endBalance = bal

		bookedStr = toDateStr(bookedAt)
		valutaStr = toDateStr(valutaDate)

		isDebit = (amt < 0)
		absAmt = q2(abs(amt))

		txs.append(BankTx(
			bookedAt=bookedStr,
			valutaDate=valutaStr,
			amount=absAmt,
			isDebit=isDebit,
			text=str(text).strip() if text is not None else '',
			balance=bal
		))
		lines += 1

	if endBalance is None:
		endBalance = q2(Decimal('0'))

	startBalance = q2(Decimal('0'))
	if txs and txs[0].balance is not None:
		first = txs[0]
		signed = (-first.amount) if first.isDebit else first.amount
		startBalance = q2(first.balance - signed)

	return {
		'startBalance': q2(startBalance),
		'endBalance': q2(endBalance),
		'txs': txs,
		'lines': lines
	}


def getBankDateRange(bankTxs: list[BankTx]) -> tuple[date, date]:
	dates: list[date] = []
	for tx in bankTxs:
		if tx.valutaDate:
			dates.append(parseDateStr(tx.valutaDate))
		elif tx.bookedAt:
			dates.append(parseDateStr(tx.bookedAt))
	if not dates:
		raise RuntimeError('No usable dates found in bank export.')
	return min(dates), max(dates)


# -----------------------------
# Matching engine
# -----------------------------
def availableDatesForAmount(byAmountMap: dict[Decimal, list[str]], countMap: dict[Key, int], amount: Decimal) -> list[str]:
	if amount not in byAmountMap:
		return []
	out: list[str] = []
	seen: set[str] = set()
	for ds in byAmountMap[amount]:
		key = (ds, amount)
		if countMap.get(key, 0) > 0 and ds not in seen:
			seen.add(ds)
			out.append(ds)
	return out


def consumeExact(countMap: dict[Key, int], amount: Decimal, dateCandidates: list[str]) -> Optional[str]:
	for ds in dateCandidates:
		key = (ds, amount)
		if countMap.get(key, 0) > 0:
			countMap[key] -= 1
			return ds
	return None


def consumeAmountOnlyIfUnique(
	countMap: dict[Key, int],
	byAmountMap: dict[Decimal, list[str]],
	amount: Decimal
) -> tuple[str, Optional[str]]:
	dates = availableDatesForAmount(byAmountMap, countMap, amount)
	if not dates:
		return 'missing', None
	if len(dates) > 1:
		return 'ambiguous', None

	used = dates[0]
	key = (used, amount)
	countMap[key] -= 1
	return 'amountOnly', used


def matchBankTxToAbacus(
	tx: BankTx,
	abacusCountMap: dict[Key, int],
	abacusByAmount: dict[Decimal, list[str]]
) -> tuple[str, Optional[str], list[str]]:
	bookedAt = None
	valutaDate = None

	if tx.bookedAt:
		try:
			bookedAt = datetime.strptime(tx.bookedAt, '%d.%m.%Y')
		except Exception:
			bookedAt = None
	if tx.valutaDate:
		try:
			valutaDate = datetime.strptime(tx.valutaDate, '%d.%m.%Y')
		except Exception:
			valutaDate = None

	candidates = buildDateCandidates(bookedAt, valutaDate, windowDays=WINDOW_DAYS)

	used = consumeExact(abacusCountMap, tx.amount, candidates)
	if used:
		return 'date', used, candidates

	mode, used2 = consumeAmountOnlyIfUnique(abacusCountMap, abacusByAmount, tx.amount)
	return mode, used2, candidates


def countLeftovers(countMap: dict[Key, int]) -> list[dict]:
	out: list[dict] = []
	for (ds, amt), cnt in countMap.items():
		if cnt > 0:
			out.append({'date': ds, 'amount': str(amt), 'count': cnt})
	out.sort(key=lambda x: (x['date'], Decimal(x['amount'])))
	return out


def buildLeftoverAmountIndex(leftovers: list[dict]) -> dict[Decimal, int]:
	idx: dict[Decimal, int] = defaultdict(int)
	for it in leftovers:
		amt = q2(Decimal(it['amount']))
		idx[amt] += int(it['count'])
	return idx


# -----------------------------
# Main
# -----------------------------
def main() -> None:
	abacusExportDir.mkdir(parents=True, exist_ok=True)
	bankExportDir.mkdir(parents=True, exist_ok=True)
	ensureLogsDir()

	abacusExports = list(abacusExportDir.glob('[!~]*.xlsx'))
	bankExports = list(bankExportDir.glob('[!~]*.xlsx'))

	if not abacusExports:
		print(f'No Abacus exports found in {abacusExportDir.resolve()}')
		return
	if not bankExports:
		print(f'No bank exports found in {bankExportDir.resolve()}')
		return

	abacusChoice = chooseFile([f.stem for f in abacusExports], 'Select Abacus account export')
	bankChoice = chooseFile([f.stem for f in bankExports], 'Select bank account export (Raiffeisen)')

	knownDiff = q2(Decimal('0'))
	if askConfirm('Known difference from previous month (that you cannot correct right now)?', default=False):
		knownDiff = askFloat('Input difference to apply on Abacus balances (example: -12.35)')

	abacusPath = Path(abacusExportDir, f'{abacusChoice}.xlsx')
	bankPath = Path(bankExportDir, f'{bankChoice}.xlsx')

	print('\nAnalyzing Raiffeisen...')
	bank = readRaiffeisen(bankPath)

	rangeMin, rangeMax = getBankDateRange(bank['txs'])
	print(f'-- Bank period detected:\t\t{rangeMin.strftime("%d.%m.%Y")} -> {rangeMax.strftime("%d.%m.%Y")}')

	print('\nAnalyzing Abacus (filtered to bank period)...')
	abacus = readAbacus(abacusPath, rangeMin, rangeMax)
	print(f'-- Abacus period used:\t\t\t{abacus["rangeMin"]} -> {abacus["rangeMax"]}')

	print('\nDocument analyze done!\n')

	linesRaiffeisen = bank['lines']
	linesAbacus = abacus['lines']

	print(f'-- Raiffeisen lines:\t{linesRaiffeisen}')
	print(f'-- Abacus lines:\t\t{linesAbacus}')

	startBalanceRaiffeisen = bank['startBalance']
	startBalanceAbacus = abacus['startBalance']
	endBalanceRaiffeisen = bank['endBalance']
	endBalanceAbacus = abacus['endBalance']

	print(f'\n-- Raiffeisen start balance:\t{startBalanceRaiffeisen}')
	print(f'-- Abacus start balance:\t\t{startBalanceAbacus}')

	print(f'\n-- Raiffeisen end balance:\t{endBalanceRaiffeisen}')
	print(f'-- Abacus end balance:\t\t{endBalanceAbacus}')

	if knownDiff != 0:
		print(f'\n-- Applying known difference to Abacus balances:\t{knownDiff}')
		startBalanceAbacus = q2(startBalanceAbacus + knownDiff)
		endBalanceAbacus = q2(endBalanceAbacus + knownDiff)

	# Matching convention:
	# - Bank credits match Abacus debits
	# - Bank debits match Abacus credits
	print('\nMatching bank -> Abacus (date-first, amount-only if unique)...')

	abacusCreditsCount = abacus['creditsCount']
	abacusDebitsCount = abacus['debitsCount']
	abacusCreditsByAmount = abacus['creditsByAmount']
	abacusDebitsByAmount = abacus['debitsByAmount']

	missingInAbacus: list[dict] = []
	dateMismatches: list[dict] = []
	ambiguousAmount: list[dict] = []

	matchedCount = 0
	matchedAmountOnly = 0

	for tx in bank['txs']:
		if tx.isDebit:
			targetCountMap = abacusCreditsCount
			targetByAmount = abacusCreditsByAmount
			targetLabel = 'Abacus credit'
		else:
			targetCountMap = abacusDebitsCount
			targetByAmount = abacusDebitsByAmount
			targetLabel = 'Abacus debit'

		mode, usedDate, _candidates = matchBankTxToAbacus(tx, targetCountMap, targetByAmount)
		prettyInfo = f'{tx.amount} CHF (booked {tx.bookedAt} / valuta {tx.valutaDate})'

		if mode == 'date':
			matchedCount += 1
			continue

		if mode == 'amountOnly':
			matchedCount += 1
			matchedAmountOnly += 1
			print(f'! DATE MISMATCH?\tBank {prettyInfo}\t-> matched {targetLabel} on Abacus {usedDate}\t[{tx.text}]')
			dateMismatches.append({
				'bankBookedAt': tx.bookedAt,
				'bankValutaDate': tx.valutaDate,
				'abacusDateUsed': usedDate,
				'amount': str(tx.amount),
				'isDebit': tx.isDebit,
				'text': tx.text
			})
			continue

		if mode == 'ambiguous':
			cands = availableDatesForAmount(targetByAmount, targetCountMap, tx.amount)
			print(f'? AMBIGUOUS\tBank {prettyInfo}\t-> amount exists multiple times in {targetLabel}, not consumed\t[{tx.text}]')
			ambiguousAmount.append({
				'bankBookedAt': tx.bookedAt,
				'bankValutaDate': tx.valutaDate,
				'amount': str(tx.amount),
				'isDebit': tx.isDebit,
				'text': tx.text,
				'candidates': cands
			})
			continue

		print(f'- MISSING\t\tBank {prettyInfo}\t-> not found in {targetLabel}\t[{tx.text}]')
		missingInAbacus.append({
			'bankBookedAt': tx.bookedAt,
			'bankValutaDate': tx.valutaDate,
			'amount': str(tx.amount),
			'isDebit': tx.isDebit,
			'text': tx.text,
			'expectedIn': targetLabel
		})

	leftoverAbacusCredits = countLeftovers(abacusCreditsCount)
	leftoverAbacusDebits = countLeftovers(abacusDebitsCount)

	# Detect inversions: missing on expected side, leftover on opposite side
	leftoverCreditsByAmount = buildLeftoverAmountIndex(leftoverAbacusCredits)
	leftoverDebitsByAmount = buildLeftoverAmountIndex(leftoverAbacusDebits)

	inversions: list[dict] = []
	for m in missingInAbacus:
		amt = q2(Decimal(m['amount']))

		# Missing bank debit expected as Abacus credit, found as leftover Abacus debit
		if m['isDebit'] and m['expectedIn'] == 'Abacus credit':
			if leftoverDebitsByAmount.get(amt, 0) > 0:
				inversions.append({
					'amount': str(amt),
					'bankBookedAt': m['bankBookedAt'],
					'bankValutaDate': m['bankValutaDate'],
					'text': m['text'],
					'expectedIn': 'Abacus credit',
					'foundAsLeftoverIn': 'Abacus debit',
					'note': 'Probable debit/credit inversion in Abacus.'
				})
				leftoverDebitsByAmount[amt] -= 1

		# Missing bank credit expected as Abacus debit, found as leftover Abacus credit
		if (not m['isDebit']) and m['expectedIn'] == 'Abacus debit':
			if leftoverCreditsByAmount.get(amt, 0) > 0:
				inversions.append({
					'amount': str(amt),
					'bankBookedAt': m['bankBookedAt'],
					'bankValutaDate': m['bankValutaDate'],
					'text': m['text'],
					'expectedIn': 'Abacus debit',
					'foundAsLeftoverIn': 'Abacus credit',
					'note': 'Probable debit/credit inversion in Abacus.'
				})
				leftoverCreditsByAmount[amt] -= 1

	# Report
	print('\n====================')
	print('FINAL REPORT')
	print('====================')
	print(f'-- Matched entries:\t\t\t{matchedCount} / {linesRaiffeisen}')
	print(f'-- Amount-only matches:\t\t{matchedAmountOnly}')
	print(f'-- Missing in Abacus:\t\t{len(missingInAbacus)}')
	print(f'-- Ambiguous amounts:\t\t{len(ambiguousAmount)}')
	print(f'-- Leftover Abacus credits:\t{sum(x["count"] for x in leftoverAbacusCredits)}')
	print(f'-- Leftover Abacus debits:\t{sum(x["count"] for x in leftoverAbacusDebits)}')

	# Balance diagnostics
	endDiff = q2(endBalanceRaiffeisen - endBalanceAbacus)

	delta = q2(Decimal('0'))
	for m in missingInAbacus:
		amt = q2(Decimal(m['amount']))
		if m['isDebit']:
			delta -= amt
		else:
			delta += amt

	print('\nBalances:')
	print(f'-- End diff (Bank - Abacus):\t{endDiff}')
	print(f'-- Missing-only delta guess:\t{delta}')
	if delta != 0:
		guess = q2(endBalanceAbacus + delta)
		print(f'-- Abacus end + delta:\t\t{guess} (target {endBalanceRaiffeisen})')

	if inversions:
		print('\nProbable inversions detected ⚠️')
		for inv in inversions:
			print(f'-- {inv["amount"]} CHF\tbank booked {inv["bankBookedAt"]} / valuta {inv["bankValutaDate"]}\t[{inv["text"]}]')
			print(f'   expected: {inv["expectedIn"]} | found leftover: {inv["foundAsLeftoverIn"]}')
			amt = q2(Decimal(inv['amount']))
			if abs(endDiff) == q2(amt * 2):
				print('   strong signal: balance diff == 2x amount')

	# Everything OK message
	if (
		len(missingInAbacus) == 0
		and len(dateMismatches) == 0
		and len(ambiguousAmount) == 0
		and sum(x["count"] for x in leftoverAbacusCredits) == 0
		and sum(x["count"] for x in leftoverAbacusDebits) == 0
		and endDiff == 0
	):
		print('\nEverything OK ✅')
		print('-- All bank transactions are matched in Abacus.')
		print('-- No date mismatches.')
		print('-- No ambiguous amounts.')
		print('-- Balances are perfectly aligned.')
	else:
		print('\nLikely causes:')
		if inversions:
			print('- At least one probable debit/credit inversion in Abacus.')
		if missingInAbacus:
			print('- Some entries look missing in Abacus (or amounts differ due to fees/rounding).')
		if ambiguousAmount:
			print('- Ambiguous repeated amounts: add tie-breaker (date preference, text/reference matching, etc.).')
		if dateMismatches:
			print('- Some entries match by amount but appear on different dates (booked vs valuta vs posting date).')

	# Logs
	dumpJson(logsDir / 'report_missingInAbacus.json', missingInAbacus)
	dumpJson(logsDir / 'report_dateMismatches.json', dateMismatches)
	dumpJson(logsDir / 'report_ambiguousAmount.json', ambiguousAmount)
	dumpJson(logsDir / 'report_leftoverAbacusCredits.json', leftoverAbacusCredits)
	dumpJson(logsDir / 'report_leftoverAbacusDebits.json', leftoverAbacusDebits)
	dumpJson(logsDir / 'report_inversions.json', inversions)

	print('\nLogs written to ./logs/')
	print('Done.')


if __name__ == '__main__':
	main()
