import json
import re
from pathlib import Path

from PyInquirer import prompt
from openpyxl import load_workbook

MATCH_DATE = re.compile(r'^[\d]{2}\.[\d]{2}\.[\d]{4}$')
#MATCH_RAIFFEISEN = re.compile(r"(?P<date>[\d]{2}\.[\d]{2}\.[\d]{2}).*?(?P<type>Crédit|Ordre|Système de recouvrement direct|Versement).*?(?P<amount>[0-9']+\.[0-9]{2}) (?P<solde>[0-9']+\.[0-9]{2})$")
#MATCH_ABACUS = re.compile(r"(?P<date>[\d]{2}\.[\d]{2}\.[\d]{4}).*?(?P<amount>[0-9']+\.[0-9]{2}) (?P<solde>[0-9']+\.[0-9]{2})")
#MATCH_ABACUS_START_SOLDE = re.compile(r"Solde y\.c\. report (?P<startSolde>[\d']+\.[\d]{2})")
#MATCH_RAIFFEISEN_START_SOLDE = re.compile(r"Report de solde (?P<startSolde>[\d']+\.[\d]{2})")
#MATCH_RAIFFEISEN_END_SOLDE = re.compile(r"Solde en votre faveur.*?(?P<endSolde>[\d']+\.[\d]{2})")

abacusExportDir = Path('abacusExports')
bankExportDir = Path('bankExports')


def parseMoney(string: str) -> float:
	try:
		return float(string.replace("'", '').strip())
	except:
		raise


def isFloat(string: str) -> bool:
	try:
		float(string)
		return True
	except:
		return False


abacusExports = abacusExportDir.glob('[!~]*.xlsx')
bankExports = bankExportDir.glob('[!~]*.xlsx')

questions = [
	{
		'type': 'list',
		'name': 'abacusFile',
		'message': 'Select abacus account export',
		'choices': [file.stem for file in abacusExports]
	},
	{
		'type': 'list',
		'name': 'bankFile',
		'message': 'Select bank account export',
		'choices': [file.stem for file in bankExports]
	},
	{
		'type': 'confirm',
		'name': 'knownDifferenceQuestion',
		'message': 'Are there any known difference with the previous month that you cannot currently correct?',
		'default': False
	},
	{
		'type': 'input',
		'name': 'knownDifference',
		'message': 'Please input the difference to apply on the Abacus end balance of the previous month',
		'validate': lambda val: isFloat(val),
		'when': lambda ans: 'knownDifferenceQuestion' in ans and ans['knownDifferenceQuestion']
	}
]

answers = prompt(questions=questions)

if not answers:
	exit(1)

abacusFile = open(Path(abacusExportDir, f'{answers["abacusFile"]}.xlsx'), mode='rb')

print('\nAnalyzing Abacus...')
workbook = load_workbook(filename=abacusFile)
endBalanceAbacus = 0
abacusCredits = dict()
abacusDebits = dict()
linesAbacus = 0
sheet = workbook.active

startBalanceAbacus = sheet['I4'].value

prevBalance = 0
for row in sheet.rows:
	row0 = str(row[0].value)
	if MATCH_DATE.match(row0):
		date = row0
		if row[6].value is not None:
			abacusDebits.setdefault(date, list())
			if float(row[6].value) in abacusDebits[date]:
				print(f'- {date} Potential debits double entry: {float(row[6].value)}')
			abacusDebits[date].append(float(row[6].value))
			linesAbacus += 1
			prevBalance = row[8].value
		elif row[7].value is not None:
			abacusCredits.setdefault(date, list())
			if float(row[7].value) in abacusCredits[date]:
				print(f'- {date} Potential credit double entry: {float(row[7].value)}')
			abacusCredits[date].append(float(row[7].value))
			linesAbacus += 1
			prevBalance = row[8].value

	elif row0.startswith('Solde') and row0 != 'Solde y.c. report':
		endBalanceAbacus = prevBalance
		break

Path('logs/abacusDebits.json').write_text(json.dumps(abacusDebits, ensure_ascii=False, indent='\t'))
Path('logs/abacusCredits.json').write_text(json.dumps(abacusCredits, ensure_ascii=False, indent='\t'))

print('\nAnalyzing Raiffeisen...')
bankCredits = dict()
bankDebits = dict()
endBalanceRaiffeisen = 0
linesRaiffeisen = 0

bankFilePath = Path(bankExportDir, f'{answers["bankFile"]}.xlsx')
bankFile = open(bankFilePath, mode='rb')
workbook = load_workbook(filename=bankFile)
sheet = workbook.active

startBalanceRaiffeisen = sheet['E2'].value - sheet['D2'].value if sheet['E2'].value > 0 else sheet['E2'].value + sheet['D2'].value
for row in sheet.rows:
	try:
		if row[0].value.lower() == 'iban':
			continue
	except:
		continue

	date = row[1].value

	try:
		date = date.strftime('%d.%m.%y')
	except:
		continue

	if row[3].value == 0:
		continue
	elif row[3].value < 1:
		bankDebits.setdefault(date, list())
		bankDebits[date].append(abs(float(row[3].value)))
		linesRaiffeisen += 1
	else:
		bankCredits.setdefault(date, list())
		bankCredits[date].append(float(row[3].value))
		linesRaiffeisen += 1

	endBalanceRaiffeisen = float(row[4].value)

Path('logs/bankDebits.json').write_text(json.dumps(bankDebits, ensure_ascii=False, indent='\t'))
Path('logs/bankCredits.json').write_text(json.dumps(bankCredits, ensure_ascii=False, indent='\t'))

print('\nDocument analyze done!')

if linesRaiffeisen > linesAbacus:
	print(f'\n- Raiffeisen has MORE entries than Abacus')
elif linesRaiffeisen < linesAbacus:
	print(f'\n- Raiffeisen has LESS entries than Abacus')

print(f'-- Raiffeisen lines: {linesRaiffeisen}')
print(f'-- Abacus lines: {linesAbacus}')

startComputedDiff = 0
if startBalanceRaiffeisen != startBalanceAbacus:
	if 'knownDifference' in answers and float(answers['knownDifference']) != 0:
		print(f'\n- Starting balances are not the same, applying known difference to Abacus start balance')
		startBalanceAbacus += float(answers['knownDifference'])
		if startBalanceRaiffeisen != startBalanceAbacus:
			print(f'-- Even after applying {answers["knownDifference"]} to Abacus start balance, balances are not matching, did you correct the month before?')
		else:
			print(f'-- After applying {answers["knownDifference"]} to Abacus start balance, the account start balance match!')
	else:
		startComputedDiff = startBalanceAbacus - startBalanceRaiffeisen
		print(f'\n- Starting balances are not the same, did you already correct the month before?')
else:
	print(f'\n- Starting balances are matching, previous months are ok!')

print(f'-- Raiffeisen start balance: {round(startBalanceRaiffeisen, 2)}')
print(f'-- Abacus start balance: {round(startBalanceAbacus, 2)}')

print(f'\n-- Raiffeisen end balance: {round(endBalanceRaiffeisen, 2)}')
print(f'-- Abacus end balance: {round(endBalanceAbacus, 2)}')

if endBalanceRaiffeisen != endBalanceAbacus:
	if 'knownDifference' in answers and float(answers['knownDifference']) != 0:
		print(f'- End balances are not the same, applying known difference to Abacus end balance')
		endBalanceAbacus += float(answers['knownDifference'])
		if endBalanceRaiffeisen != endBalanceAbacus:
			print(f'- Even after applying {answers["knownDifference"]} to Abacus end balance, balances are not matching, you got work to do!')
		else:
			print(f'- After applying {answers["knownDifference"]} to Abacus end balance, the account balance match!')
	else:
		print(f'\n- Ending balances are not the same')
		if startComputedDiff != 0:
			if round(endBalanceAbacus - startComputedDiff, 2) == round(endBalanceRaiffeisen, 2):
				print(f'-- If you correct the starting balance, meaning correct the past months and find the {startComputedDiff} difference, the end balance would match')
			else:
				print(f'-- Even after applying the start difference of {startComputedDiff}, the balance don\'t match, something is wrong')
else:
	print(f'\n- Ending balances are matching, so far so good!')



print('\nMatching Raiffeisen credits to Abacus debits...')

notFoundBankCredits = list()
potentialEndBalanceCorrection = 0
missingCreditsCount = 0
for date, amounts in bankCredits.copy().items():
	for amount in amounts:
		found = False
		for abacusDate, abacusAmounts in abacusDebits.copy().items():
			for abacusAmount in abacusAmounts.copy():
				if amount == abacusAmount:
					found = True
					abacusDebits[abacusDate].remove(abacusAmount)
					break
			if found:
				break
		if not found:
			print(f'- Raiffeisen credit of {amount} CHF on {date} not found in Abacus')
			missingCreditsCount += 1
			potentialEndBalanceCorrection += amount
			notFoundBankCredits.append(amount)

if missingCreditsCount > 0:
	print(f'\n{missingCreditsCount} missing credits in Abacus')

if potentialEndBalanceCorrection > 0:
	print(f'\n-- Applying potential correction, new end balance on Abacus: {round(endBalanceAbacus + potentialEndBalanceCorrection)}')


print('\nMatching Raiffeisen debits to Abacus credits...')
notFoundBankDebits = list()
missingDebitsCount = 0
for date, amounts in bankDebits.copy().items():
	for amount in amounts:
		found = False
		for abacusDate, abacusAmounts in abacusCredits.copy().items():
			for abacusAmount in abacusAmounts.copy():
				if amount == abacusAmount:
					found = True
					abacusCredits[abacusDate].remove(amount)
					break
			if found:
				break
		if not found:
			print(f'- Raiffeisen dedit of {amount} CHF on {date} not found in Abacus')
			missingDebitsCount += 1
			potentialEndBalanceCorrection -= amount
			notFoundBankDebits.append(amount)

if missingDebitsCount > 0:
	print(f'\n{missingDebitsCount} missing debits in Abacus')

newPotentialAbacusBalance = endBalanceAbacus
if potentialEndBalanceCorrection != 0:
	newPotentialAbacusBalance = round(endBalanceAbacus + potentialEndBalanceCorrection, 2)
	print(f'\n-- Applying potential correction, new end balance on Abacus: {newPotentialAbacusBalance}')
	if newPotentialAbacusBalance == endBalanceRaiffeisen:
		print('--- Correcting missing parts would fix the account and bring it to same level as the bank!')
	else:
		print("--- Applying corrections does not fix the account, there's something else...")

print('\nThese entries are not matched to Raiffeisen entries')
correction = 0
inverseAbacusCredits = list()
inverseAbacusDebits = list()
removeFromAbacus = list()
print('- Credits')
for date, listing in abacusCredits.items():
	for amount in listing:
		print(f'-- {date} {amount}')
		correction += amount
		if amount in notFoundBankCredits:
			inverseAbacusCredits.append(str(amount))
		else:
			removeFromAbacus.append(str(amount))

print('- Debits')
for date, listing in abacusDebits.items():
	for amount in listing:
		print(f'-- {date} {amount}')
		correction -= amount
		if amount in notFoundBankDebits:
			inverseAbacusDebits.append(str(amount))
		else:
			removeFromAbacus.append(str(amount))

if correction != 0:
	print('\n-- Adding non matched credits and removing non matched debits...')
	result = round(newPotentialAbacusBalance + correction, 2)
	print(f'--- Calculated result on Abacus account: {result} {f"(should be {endBalanceRaiffeisen})" if result != endBalanceRaiffeisen else "and that IS A MATCH, we rock!"}')
	if result == endBalanceRaiffeisen:
		print('---- To fix the balances you should:')
		if missingDebitsCount > 0:
			print(f'----- Apply the {missingDebitsCount} missing debits')

		if missingCreditsCount > 0:
			print(f'----- Apply the {missingCreditsCount} missing credits')

		if inverseAbacusCredits:
			print(f'----- Inverse the following entrie(s) to match the accounts: {", ".join(inverseAbacusCredits)}')

		if inverseAbacusDebits:
			print(f'----- Inverse the following entrie(s) to match the accounts: {", ".join(inverseAbacusDebits)}')

		if removeFromAbacus:
			print(f'----- Remove the following entrie(s) to match the accounts: {", ".join(removeFromAbacus)}')
	else:
		print('--- Accounts not matching, something else is wrong...')
