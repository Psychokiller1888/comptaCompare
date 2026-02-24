import json
import re
from datetime import datetime
from pathlib import Path

from PyInquirer import prompt
from openpyxl import load_workbook

MATCH_DATE = re.compile(r'^[\d]{2}\.[\d]{2}\.[\d]{4}$')

abacusExportDir = Path('abacusExports')
bankExportDir = Path('bankExports')
logsDir = Path('logs')


def ensureDirs() -> None:
	abacusExportDir.mkdir(parents=True, exist_ok=True)
	bankExportDir.mkdir(parents=True, exist_ok=True)
	logsDir.mkdir(parents=True, exist_ok=True)


def parseMoney(string: str) -> float:
	# support "123'456.78" and "123456.78" and "123456,78"
	s = str(string).replace("'", '').replace(' ', '').strip()
	s = s.replace(',', '.')
	return float(s)


def isFloat(string: str) -> bool:
	try:
		parseMoney(string)
		return True
	except Exception:
		return False


def safeRound(value) -> float:
	# round floats and numeric strings safely
	return round(parseMoney(value), 2)


ensureDirs()

abacusExports = list(abacusExportDir.glob('[!~]*.xlsx'))
bankExports = list(bankExportDir.glob('[!~]*.xlsx'))

if not abacusExports:
	print('No Abacus exports found in abacusExports/')
	exit(1)

if not bankExports:
	print('No bank exports found in bankExports/')
	exit(1)

questions = [
	{
		'type': 'list',
		'name': 'abacusFile',
		'message': 'Select abacus account export',
		'choices': [file.stem for file in abacusExports]
	},
	{
		'type': 'list',
		'name': 'bankFile',
		'message': 'Select bank account export',
		'choices': [file.stem for file in bankExports]
	},
	{
		'type': 'confirm',
		'name': 'knownDifferenceQuestion',
		'message': 'Are there any known difference with the previous month that you cannot currently correct?',
		'default': False
	},
	{
		'type': 'input',
		'name': 'knownDifference',
		'message': 'Please input the difference to apply on the Abacus end balance of the previous month',
		'validate': lambda val: isFloat(val),
		'when': lambda ans: 'knownDifferenceQuestion' in ans and ans['knownDifferenceQuestion']
	}
]

answers = prompt(questions=questions)

if not answers:
	exit(1)

knownDifference = 0.0
if 'knownDifference' in answers and answers['knownDifference'] is not None:
	knownDifference = parseMoney(answers['knownDifference'])

# -------------------------
# Analyzing Abacus
# -------------------------
abacusPath = Path(abacusExportDir, f'{answers["abacusFile"]}.xlsx')

print('\nAnalyzing Abacus...')
workbook = load_workbook(filename=str(abacusPath), data_only=True)
sheet = workbook.active

endBalanceAbacus = 0.0
abacusCredits = dict()
abacusDebits = dict()
linesAbacus = 0

startBalanceAbacus = sheet['I4'].value
if startBalanceAbacus is None:
	startBalanceAbacus = 0.0
else:
	startBalanceAbacus = safeRound(startBalanceAbacus)

prevBalance = 0.0

for row in sheet.rows:
	row0 = '' if row[0].value is None else str(row[0].value)

	if MATCH_DATE.match(row0):
		date = row0
		debit = row[6].value
		credit = row[7].value

		# On garde exactement ta logique:
		# - si debit pas None => debit
		# - elif credit pas None => credit
		if debit is not None:
			debit = safeRound(debit)
			abacusDebits.setdefault(date, list())
			if debit in abacusDebits[date]:
				print(f'- {date} Potential debits double entry: {debit}')
			abacusDebits[date].append(debit)
			linesAbacus += 1
			if row[8].value is not None:
				prevBalance = safeRound(row[8].value)

		elif credit is not None:
			credit = safeRound(credit)
			abacusCredits.setdefault(date, list())
			if credit in abacusCredits[date]:
				print(f'- {date} Potential credit double entry: {credit}')
			abacusCredits[date].append(credit)
			linesAbacus += 1
			if row[8].value is not None:
				prevBalance = safeRound(row[8].value)

	elif row0.startswith('Solde') and row0 != 'Solde y.c. report':
		endBalanceAbacus = prevBalance
		break

endBalanceAbacus = safeRound(endBalanceAbacus)

(logsDir / 'abacusDebits.json').write_text(json.dumps(abacusDebits, ensure_ascii=False, indent='\t'))
(logsDir / 'abacusCredits.json').write_text(json.dumps(abacusCredits, ensure_ascii=False, indent='\t'))

# -------------------------
# Analyzing Raiffeisen
# -------------------------
print('\nAnalyzing Raiffeisen...')
bankCredits = dict()
bankDebits = dict()
endBalanceRaiffeisen = 0.0
linesRaiffeisen = 0

bankFilePath = Path(bankExportDir, f'{answers["bankFile"]}.xlsx')
workbook = load_workbook(filename=str(bankFilePath), data_only=True)
sheet = workbook.active

# Ta formule, mais safe:
e2 = sheet['E2'].value or 0
d2 = sheet['D2'].value or 0
e2 = parseMoney(e2)
d2 = parseMoney(d2)

startBalanceRaiffeisen = e2 - d2 if e2 > 0 else e2 + d2
startBalanceRaiffeisen = safeRound(startBalanceRaiffeisen)

for row in sheet.rows:
	try:
		if row[0].value and str(row[0].value).lower() == 'iban':
			continue
	except Exception:
		continue

	try:
		date = row[1].value.strftime('%d.%m.%Y')
	except Exception as e:
		# On garde ton print d'erreur
		print(f'Error: {e}')
		continue

	try:
		data = safeRound(row[3].value)
	except Exception:
		continue

	if data == 0:
		continue
	elif data < 0:
		bankDebits.setdefault(date, list())
		bankDebits[date].append(abs(data))
		linesRaiffeisen += 1
	else:
		bankCredits.setdefault(date, list())
		bankCredits[date].append(data)
		linesRaiffeisen += 1

	# colonne solde banque (row[4]) comme tu fais
	try:
		endBalanceRaiffeisen = safeRound(row[4].value)
	except Exception:
		pass

(logsDir / 'bankDebits.json').write_text(json.dumps(bankDebits, ensure_ascii=False, indent='\t'))
(logsDir / 'bankCredits.json').write_text(json.dumps(bankCredits, ensure_ascii=False, indent='\t'))

print('\nDocument analyze done!')

if linesRaiffeisen > linesAbacus:
	print(f'\n- Raiffeisen has MORE entries than Abacus')
elif linesRaiffeisen < linesAbacus:
	print(f'\n- Raiffeisen has LESS entries than Abacus')

print(f'-- Raiffeisen lines: {linesRaiffeisen}')
print(f'-- Abacus lines: {linesAbacus}')

startComputedDiff = 0.0
if startBalanceRaiffeisen != startBalanceAbacus:
	if knownDifference != 0:
		print(f'\n- Starting balances are not the same, applying known difference to Abacus start balance')
		startBalanceAbacus = safeRound(startBalanceAbacus + knownDifference)
		if startBalanceRaiffeisen != startBalanceAbacus:
			print(f'-- Even after applying {knownDifference} to Abacus start balance, balances are not matching, did you correct the month before?')
		else:
			print(f'-- After applying {knownDifference} to Abacus start balance, the account start balance match!')
	else:
		startComputedDiff = startBalanceAbacus - startBalanceRaiffeisen
		print(f'\n- Starting balances are not the same, did you already correct the month before?')
else:
	print(f'\n- Starting balances are matching, previous months are ok!')

print(f'-- Raiffeisen start balance: {startBalanceRaiffeisen}')
print(f'-- Abacus start balance: {startBalanceAbacus}')

print(f'\n-- Raiffeisen end balance: {endBalanceRaiffeisen}')
print(f'-- Abacus end balance: {endBalanceAbacus}')

if endBalanceRaiffeisen != endBalanceAbacus:
	if knownDifference != 0:
		print(f'-- End balances are not the same, applying known difference to Abacus end balance')
		endBalanceAbacus = safeRound(endBalanceAbacus + knownDifference)
		if endBalanceRaiffeisen != endBalanceAbacus:
			print(f'--- Even after applying {knownDifference} to Abacus end balance, balances are not matching, you got work to do!')
			print(f'---- Recalculated Abacus end balance {endBalanceAbacus}, Raiffeisen end balance {endBalanceRaiffeisen}')
		else:
			print(f'--- After applying {knownDifference} to Abacus end balance, the account balance match, so we\'re all good!')
			exit(0)
	else:
		print(f'\n- Ending balances are not the same')
		if startComputedDiff != 0:
			if safeRound(endBalanceAbacus - startComputedDiff) == endBalanceRaiffeisen:
				print(f'-- If you correct the starting balance, meaning correct the past months and find the {safeRound(startComputedDiff)} difference, the end balance would match')
			else:
				print(f'-- Even after applying the start difference of {startComputedDiff}, the balance don\'t match, something is wrong')
else:
	print(f'\n- Ending balances are matching, so far so good!')

# -------------------------
# Matching credits/debits
# -------------------------
print('\nMatching Raiffeisen credits to Abacus debits...')

notFoundBankCredits = list()
potentialEndBalanceCorrection = 0.0
missingCreditsCount = 0

for date, amounts in bankCredits.copy().items():
	for amount in amounts:
		found = False

		# Match by date ONLY (default)
		if date in abacusDebits:
			for abacusAmount in abacusDebits[date].copy():
				if amount == abacusAmount:
					found = True
					abacusDebits[date].remove(abacusAmount)
					break

		if not found:
			print(f'- Raiffeisen credit of {amount} CHF on {date} not found in Abacus')
			missingCreditsCount += 1
			potentialEndBalanceCorrection += amount
			notFoundBankCredits.append(amount)

if missingCreditsCount > 0:
	print(f'\n{missingCreditsCount} missing credits in Abacus')

if potentialEndBalanceCorrection > 0:
	print(f'\n-- Applying potential correction, new end balance on Abacus: {round(endBalanceAbacus + potentialEndBalanceCorrection)}')

print('\nMatching Raiffeisen debits to Abacus credits...')
notFoundBankDebits = list()
missingDebitsCount = 0

for date, amounts in bankDebits.copy().items():
	for amount in amounts:
		found = False

		# Match by date ONLY (default)
		if date in abacusCredits:
			for abacusAmount in abacusCredits[date].copy():
				if amount == abacusAmount:
					found = True
					abacusCredits[date].remove(abacusAmount)
					break

		if not found:
			print(f'- Raiffeisen debit of {amount} CHF on {date} not found in Abacus')
			missingDebitsCount += 1
			potentialEndBalanceCorrection -= amount
			notFoundBankDebits.append(amount)

if missingDebitsCount > 0:
	print(f'\n{missingDebitsCount} missing debits in Abacus')

newPotentialAbacusBalance = endBalanceAbacus
if potentialEndBalanceCorrection != 0:
	newPotentialAbacusBalance = safeRound(endBalanceAbacus + potentialEndBalanceCorrection)
	print(f'\n-- Applying potential correction, new end balance on Abacus: {newPotentialAbacusBalance}')
	if newPotentialAbacusBalance == endBalanceRaiffeisen:
		print('--- Correcting missing parts would fix the account and bring it to same level as the bank!')
	else:
		print('--- Applying corrections does not fix the account, there\'s something else...')

print('\nThese entries are not matched to Raiffeisen entries')
correction = 0.0
inverseAbacusCredits = list()
inverseAbacusDebits = list()
removeFromAbacus = list()

print('- Credits')
for date, listing in abacusCredits.items():
	for amount in listing:
		print(f'-- {date} {amount}')
		correction += amount
		if amount in notFoundBankCredits:
			inverseAbacusCredits.append(str(amount))
		else:
			removeFromAbacus.append(str(amount))

print('- Debits')
for date, listing in abacusDebits.items():
	for amount in listing:
		print(f'-- {date} {amount}')
		correction -= amount
		if amount in notFoundBankDebits:
			inverseAbacusDebits.append(str(amount))
		else:
			removeFromAbacus.append(str(amount))

if correction != 0:
	print('\n-- Adding non matched credits and removing non matched debits...')
	result = safeRound(newPotentialAbacusBalance + correction)
	if result != safeRound(endBalanceRaiffeisen):
		print(f'--- Calculated result on Abacus account: {result} (should be {safeRound(endBalanceRaiffeisen)}, difference {safeRound(result - endBalanceRaiffeisen)})')
	else:
		print(f'--- Calculated result on Abacus account: {result} and that IS A MATCH, we rock!')

	if result == safeRound(endBalanceRaiffeisen):
		print('---- To fix the balances you should:')
		if missingDebitsCount > 0:
			print(f'----- Apply the {missingDebitsCount} missing debits')

		if missingCreditsCount > 0:
			print(f'----- Apply the {missingCreditsCount} missing credits')

		if inverseAbacusCredits:
			print(f'----- Inverse the following entrie(s) to match the accounts: {", ".join(inverseAbacusCredits)}')

		if inverseAbacusDebits:
			print(f'----- Inverse the following entrie(s) to match the accounts: {", ".join(inverseAbacusDebits)}')

		if removeFromAbacus:
			print(f'----- Remove the following entrie(s) to match the accounts: {", ".join(removeFromAbacus)}')
	else:
		print('--- Accounts not matching, something else is wrong...')
